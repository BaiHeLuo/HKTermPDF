"""
批量生成「輸入勞工終止僱傭合約通知書」PDF - 核心逻辑
通过表头名称查找Excel列，而非硬编码列位置。
"""

import os
import openpyxl
import fitz  # PyMuPDF


# 需要从Excel中读取的字段及其对应的表头关键字（模糊匹配）
COLUMN_MAPPING = {
    "中文姓名": ["中文姓名"],
    "英文姓名": ["英文姓名"],
    "約滿日期": ["約滿日期"],
    "入境事務處檔號": ["入境事務處檔號"],
    "香港身份證號碼": ["香港身份證號碼"],
    "ESLS": ["SLS"],          # 表头可能是 "SLS" 或 "ESLS"
    "勞工處檔案編號": ["勞工處檔案編號", "勞工處配額編號"],
}


def normalize(text):
    """去除空格，用于模糊匹配表头"""
    return str(text).replace(" ", "").replace("\u3000", "").strip()


def find_columns(ws):
    """
    在Excel工作表中查找表头行，并按名称映射列索引。
    返回 (header_row_index, {field_name: col_index})
    """
    # 遍历前10行找表头
    for row_idx in range(1, min(ws.max_row + 1, 11)):
        row_values = [normalize(cell.value) if cell.value else "" for cell in ws[row_idx]]

        # 检查这一行是否包含关键表头
        if any("中文姓名" in v for v in row_values):
            col_map = {}
            for field_name, keywords in COLUMN_MAPPING.items():
                for kw in keywords:
                    for col_idx, val in enumerate(row_values):
                        if kw in val or val in kw:
                            col_map[field_name] = col_idx
                            break
                    if field_name in col_map:
                        break

            # 验证必须字段
            required = ["中文姓名", "英文姓名", "約滿日期", "入境事務處檔號",
                        "香港身份證號碼", "ESLS", "勞工處檔案編號"]
            missing = [r for r in required if r not in col_map]
            if missing:
                raise ValueError(f"Excel表头缺少以下字段: {', '.join(missing)}")

            return row_idx, col_map

    raise ValueError("未在Excel前10行中找到包含'中文姓名'的表头行")


def parse_date(date_str):
    """解析日期字符串，如 '2024.06.20' → (dd, mm, yyyy)"""
    parts = str(date_str).strip().split(".")
    if len(parts) == 3:
        yyyy, mm, dd = parts[0], parts[1], parts[2]
        return dd, mm, yyyy
    raise ValueError(f"无法解析日期: {date_str}，期望格式如 2024.06.20")


def extract_esls_number(esls_str):
    """从 ESLS004030 提取 004030（去掉 ESLS 前缀）"""
    s = str(esls_str).strip()
    if s.upper().startswith("ESLS"):
        return s[4:]
    return s


def read_workers_from_excel(filepath):
    """从Excel读取勞工数据，通过表头名称查找列"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header_row, col_map = find_columns(ws)

    workers = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
        cn_name_col = col_map["中文姓名"]
        if not row[cn_name_col].value:
            continue

        worker = {"_row": row_idx}  # 记录行号，方便报错
        for field_name, col_idx in col_map.items():
            worker[field_name] = row[col_idx].value

        workers.append(worker)

    return workers


def validate_worker(worker):
    """
    验证单个勞工数据，返回错误列表。
    每条错误格式: "第X行 - 姓名: 错误描述"
    """
    errors = []
    row = worker.get("_row", "?")
    name = worker.get("中文姓名", "未知")
    prefix = f"第{row}行 [{name}]"

    # 检查必填字段是否为空
    required_fields = {
        "中文姓名": "中文姓名",
        "英文姓名": "英文姓名",
        "約滿日期": "約滿日期",
        "入境事務處檔號": "入境事務處檔號",
        "香港身份證號碼": "香港身份證號碼",
        "ESLS": "SLS（僱傭合約號碼）",
        "勞工處檔案編號": "勞工處檔案編號",
    }
    for field_key, display_name in required_fields.items():
        val = worker.get(field_key)
        if val is None or str(val).strip() == "":
            errors.append(f"{prefix}: 缺少「{display_name}」")

    # 验证日期格式
    date_val = worker.get("約滿日期")
    if date_val:
        parts = str(date_val).strip().split(".")
        if len(parts) != 3:
            errors.append(f"{prefix}: 約滿日期格式錯誤「{date_val}」，應為 yyyy.mm.dd（如 2026.06.20）")
        else:
            try:
                yyyy, mm, dd = int(parts[0]), int(parts[1]), int(parts[2])
                if not (1900 <= yyyy <= 2100):
                    errors.append(f"{prefix}: 約滿日期年份異常「{yyyy}」")
                if not (1 <= mm <= 12):
                    errors.append(f"{prefix}: 約滿日期月份異常「{mm}」")
                if not (1 <= dd <= 31):
                    errors.append(f"{prefix}: 約滿日期日份異常「{dd}」")
            except ValueError:
                errors.append(f"{prefix}: 約滿日期無法解析「{date_val}」，應為 yyyy.mm.dd")

    # 验证姓名是否含有异常字符
    cn_name = str(worker.get("中文姓名", "")).strip()
    en_name = str(worker.get("英文姓名", "")).strip()
    if cn_name and len(cn_name) < 2:
        errors.append(f"{prefix}: 中文姓名過短「{cn_name}」，請確認是否正確")
    if en_name and not any(c.isalpha() for c in en_name):
        errors.append(f"{prefix}: 英文姓名「{en_name}」不含字母，請確認是否正確")

    return errors


def fill_pdf(template_path, worker, output_path, signing_date, phone_no="24367886",
             termination_reason="劳工合约到期"):
    """
    填写PDF表单

    Args:
        template_path: 模板PDF路径
        worker: 勞工数据字典
        output_path: 输出路径
        signing_date: 签署日期字符串 (格式: d/m/yyyy)
        phone_no: 联系电话
        termination_reason: 终止合约原因
    """
    doc = fitz.open(template_path)

    dd, mm, yyyy = parse_date(worker["約滿日期"])
    esls_no = extract_esls_number(worker["ESLS"])

    # ===== 第一页 =====
    for widget in doc[0].widgets():
        name = widget.field_name
        if name == "僱傭合約號碼 Employment Contract No":
            widget.field_value = esls_no
            widget.update()
        elif name == "勞工處檔案編號 LD Reference No":
            widget.field_value = str(worker["勞工處檔案編號"])
            widget.update()
        elif name == "入境事務處簽證／進入許可申請檔號編號 ImmD Application Reference No. for Visa/ Entry Permit":
            widget.field_value = str(worker["入境事務處檔號"])
            widget.update()

    # ===== 第二页 =====
    for widget in doc[1].widgets():
        name = widget.field_name
        if name == "姓名":
            widget.field_value = worker["中文姓名"]
            widget.update()
        elif name == "Name":
            widget.field_value = worker["英文姓名"]
            widget.update()
        elif name == "香港身份證號碼 HKIC No":
            widget.field_value = str(worker["香港身份證號碼"])
            widget.update()
        elif name == "輸入勞工聯絡電話 Phone No":
            widget.field_value = phone_no
            widget.update()
        elif name == "日 dd":
            widget.field_value = dd
            widget.update()
        elif name == "月 mm":
            widget.field_value = mm
            widget.update()
        elif name == "年 yyyy":
            widget.field_value = yyyy
            widget.update()
        elif name == "終止合約原因 Reason(s) for termination of contract":
            widget.field_value = termination_reason
            widget.update()
        elif name == "日期 Date_af_date":
            widget.field_value = signing_date
            widget.update()

    doc.save(output_path)
    doc.close()
